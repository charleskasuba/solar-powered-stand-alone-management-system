import requests
import json
import time
import threading
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from flask import Flask, render_template_string, jsonify, send_file, request
from flask_cors import CORS
import math
import os
import csv
import sqlite3
import calendar
from collections import deque
from io import BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import warnings
warnings.filterwarnings('ignore')

app = Flask(__name__)
CORS(app)

# ========== CONFIGURATION ==========
COPPERBELT_LAT = -12.82
COPPERBELT_LON = 28.21
HISTORICAL_DATA_FILE = 'historical_power_data.csv'
DB_FILE = 'solar_data.db'
ZMW_PER_KWH = 2.5
KG_CO2_PER_KWH = 0.5
BIRD_API_KEY = "bk_eu1_CX53rclKcsQZUy5Hat1KAnC4S18uB"
EMAIL_TO = "mwansakapikila48@gmail.com"

# ========== GLOBAL DATA STORAGE ==========
system_data = {
    'battery_voltage': 0, 'battery_current': 0, 'battery_soc': 0,
    'inverter_voltage': 0, 'inverter_current': 0, 'inverter_power': 0,
    'load_voltage': 0, 'load_current': 0, 'load_power': 0,
    'energy_consumed_kwh': 0, 'energy_produced_kwh': 0,
    'load1_state': 'OFF', 'load2_state': 'OFF',
    'load1_power': 0, 'load2_power': 0,
    'trip_state': 'NOR', 'power_balance': 0, 'phase': 'P1',
    'weather': {'current': {'temperature': 25, 'condition': 'Mild'}, 'daily': []},
    'sunlight': {'uv_index': 5, 'sunrise': '06:00', 'sunset': '18:00'},
    'predictions': {'today_energy': 0, 'tomorrow_energy': 0, 'weekly_energy': 0, 'peak_hours': []},
    'recommendations': [],
    'esp32_online': False, 'last_esp32_seen': None,
    'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
}

command_queue = []
command_lock = threading.Lock()
real_time_power = deque(maxlen=60)
historical_power = []

# ========== EMAIL NOTIFICATIONS ==========
def send_email(subject, html_body):
    try:
        url = "https://eu1.platform.bird.com/v1/email/messages"
        headers = {
            "Authorization": f"Bearer {BIRD_API_KEY}",
            "Content-Type": "application/json"
        }
        payload = {
            "from": "onboarding@messagebird.dev",
            "to": [EMAIL_TO],
            "subject": subject,
            "html": html_body
        }
        response = requests.post(url, json=payload, headers=headers, timeout=15)
        if response.status_code in (200, 201, 202):
            print(f"Email sent: {subject}")
            return True
        else:
            print(f"Email error {response.status_code}: {response.text[:300]}")
            return False
    except Exception as e:
        print(f"Email send error: {e}")
        return False

def notify_esp32_offline():
    html = """
    <div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;padding:20px">
        <div style="background:#f44336;color:white;padding:15px;border-radius:8px;text-align:center">
            <h2>Solar Microgrid Alert</h2>
        </div>
        <div style="padding:20px;background:#fff3f3;border:1px solid #f44336;border-radius:0 0 8px 8px">
            <h3 style="color:#f44336">ESP32 is OFFLINE</h3>
            <p>The ESP32 controller has lost connection with the server.</p>
            <p><strong>Last seen:</strong> """ + str(system_data['last_esp32_seen'] or 'Never') + """</p>
            <p><strong>Time:</strong> """ + datetime.now().strftime('%Y-%m-%d %H:%M:%S') + """</p>
            <hr>
            <p style="color:#666;font-size:12px">Copperbelt University Solar Microgrid System</p>
        </div>
    </div>
    """
    send_email("Solar Microgrid Alert: ESP32 OFFLINE", html)

def notify_low_battery(soc):
    html = """
    <div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;padding:20px">
        <div style="background:#ff9800;color:white;padding:15px;border-radius:8px;text-align:center">
            <h2>Solar Microgrid Warning</h2>
        </div>
        <div style="padding:20px;background:#fff8e1;border:1px solid #ff9800;border-radius:0 0 8px 8px">
            <h3 style="color:#ff9800">Low Battery Warning</h3>
            <p>Battery State of Charge has dropped below safe levels.</p>
            <p><strong>Current SOC:</strong> """ + f"{soc:.1f}%" + """</p>
            <p><strong>Voltage:</strong> """ + f"{system_data['battery_voltage']:.2f}V" + """</p>
            <p><strong>Load Power:</strong> """ + f"{system_data['load_power']:.1f}W" + """</p>
            <p><strong>Time:</strong> """ + datetime.now().strftime('%Y-%m-%d %H:%M:%S') + """</p>
            <hr>
            <p style="color:#666;font-size:12px">Copperbelt University Solar Microgrid System</p>
        </div>
    </div>
    """
    send_email(f"Solar Microgrid: Low Battery ({soc:.0f}%)", html)

def notify_esp32_online():
    html = """
    <div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;padding:20px">
        <div style="background:#4CAF50;color:white;padding:15px;border-radius:8px;text-align:center">
            <h2>Solar Microgrid Status</h2>
        </div>
        <div style="padding:20px;background:#f1f8e9;border:1px solid #4CAF50;border-radius:0 0 8px 8px">
            <h3 style="color:#4CAF50">ESP32 is BACK ONLINE</h3>
            <p>The ESP32 controller has reconnected successfully.</p>
            <p><strong>Battery SOC:</strong> """ + f"{system_data['battery_soc']:.1f}%" + """</p>
            <p><strong>Time:</strong> """ + datetime.now().strftime('%Y-%m-%d %H:%M:%S') + """</p>
            <hr>
            <p style="color:#666;font-size:12px">Copperbelt University Solar Microgrid System</p>
        </div>
    </div>
    """
    send_email("Solar Microgrid: ESP32 Back Online", html)

# Track email notification state to avoid spam
email_alert_state = {'esp32_was_online': False, 'low_battery_sent': False}

# ========== DATABASE ==========
def get_db():
    conn = sqlite3.connect(DB_FILE, timeout=10)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    try:
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute('''CREATE TABLE IF NOT EXISTS sensor_readings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp TEXT DEFAULT CURRENT_TIMESTAMP,
            battery_voltage REAL, battery_current REAL, battery_soc REAL,
            inverter_voltage REAL, inverter_current REAL, inverter_power REAL,
            load_voltage REAL, load_current REAL, load_power REAL,
            energy_consumed_kwh REAL, energy_produced_kwh REAL,
            relay1_state TEXT, relay2_state TEXT, power_balance REAL
        )''')
        c.execute('''CREATE TABLE IF NOT EXISTS relay_events (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp TEXT DEFAULT CURRENT_TIMESTAMP,
            relay INTEGER, state TEXT, source TEXT
        )''')
        conn.commit()
        conn.close()
        print("Database initialized")
    except Exception as e:
        print(f"DB init error: {e}")

def log_sensor_reading():
    try:
        conn = sqlite3.connect(DB_FILE, timeout=10)
        c = conn.cursor()
        c.execute('''INSERT INTO sensor_readings
            (timestamp, battery_voltage, battery_current, battery_soc,
             inverter_voltage, inverter_current, inverter_power,
             load_voltage, load_current, load_power,
             energy_consumed_kwh, energy_produced_kwh,
             relay1_state, relay2_state, power_balance)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
            (datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
             system_data['battery_voltage'], system_data['battery_current'], system_data['battery_soc'],
             system_data['inverter_voltage'], system_data['inverter_current'], system_data['inverter_power'],
             system_data['load_voltage'], system_data['load_current'], system_data['load_power'],
             system_data['energy_consumed_kwh'], system_data['energy_produced_kwh'],
             system_data['load1_state'], system_data['load2_state'], system_data['power_balance']))
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"DB log error: {e}")

def log_relay_event(relay, state, source='manual'):
    try:
        conn = sqlite3.connect(DB_FILE, timeout=10)
        c = conn.cursor()
        c.execute('INSERT INTO relay_events (timestamp, relay, state, source) VALUES (?, ?, ?, ?)',
                  (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), relay, state, source))
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"DB relay log error: {e}")

def get_readings_between(start_date, end_date):
    try:
        conn = get_db()
        c = conn.cursor()
        c.execute('SELECT * FROM sensor_readings WHERE timestamp >= ? AND timestamp <= ? ORDER BY timestamp',
                  (start_date, end_date))
        rows = [dict(r) for r in c.fetchall()]
        conn.close()
        return rows
    except Exception as e:
        print(f"DB read error: {e}")
        return []

def get_relay_events_between(start_date, end_date):
    try:
        conn = get_db()
        c = conn.cursor()
        c.execute('SELECT * FROM relay_events WHERE timestamp >= ? AND timestamp <= ? ORDER BY timestamp',
                  (start_date, end_date))
        rows = [dict(r) for r in c.fetchall()]
        conn.close()
        return rows
    except Exception as e:
        print(f"DB relay read error: {e}")
        return []

# ========== HISTORICAL DATA ==========
def init_historical_data():
    global historical_power
    print("Generating 2 months of historical data...")
    end_date = datetime.now()
    start_date = end_date - timedelta(days=60)
    current_date = start_date
    while current_date <= end_date:
        hour = current_date.hour
        is_weekend = current_date.weekday() >= 5
        if 6 <= hour < 9:
            consumed = 388 + np.random.normal(0, 30)
            produced = 200 + np.random.normal(0, 50)
        elif 9 <= hour < 12:
            consumed = 554 + np.random.normal(0, 40)
            produced = 500 + np.random.normal(0, 80)
        elif 12 <= hour < 14:
            consumed = 554 + np.random.normal(0, 50)
            produced = 700 + np.random.normal(0, 60)
        elif 14 <= hour < 17:
            consumed = 554 + np.random.normal(0, 40)
            produced = 600 + np.random.normal(0, 70)
        elif 17 <= hour < 21:
            consumed = 752 + np.random.normal(0, 60)
            produced = 300 + np.random.normal(0, 50)
        else:
            consumed = 200 + np.random.normal(0, 20)
            produced = 0
        if is_weekend:
            consumed = consumed * 1.15
        historical_power.append({
            'timestamp': current_date,
            'consumed': max(0, consumed),
            'produced': max(0, produced)
        })
        current_date += timedelta(hours=1)
    save_historical_to_csv()
    print(f"Generated {len(historical_power)} historical records")

def save_historical_to_csv():
    try:
        with open(HISTORICAL_DATA_FILE, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['timestamp', 'consumed_power', 'produced_power'])
            for record in historical_power:
                writer.writerow([record['timestamp'], record['consumed'], record['produced']])
    except Exception as e:
        print(f"CSV save error: {e}")

# ========== PREDICTIONS ==========
def calculate_energy_predictions():
    if len(historical_power) < 24:
        return {'today_energy': 5.5, 'tomorrow_energy': 5.8, 'weekly_energy': 40, 'peak_hours': []}
    try:
        df = pd.DataFrame(historical_power)
        df['hour'] = df['timestamp'].dt.hour
        hourly_avg = df.groupby('hour')['consumed'].mean()
        current_hour = datetime.now().hour
        remaining_hours = 24 - current_hour
        current_avg = hourly_avg.get(current_hour, 500)
        today_energy = (current_avg * remaining_hours + system_data['energy_consumed_kwh'] * 1000) / 1000
        peak_hours = []
        threshold = hourly_avg.quantile(0.75)
        for hour in range(24):
            if hourly_avg[hour] > threshold:
                peak_hours.append({
                    'hour': f"{hour:02d}:00-{(hour+1):02d}:00",
                    'factor': round(hourly_avg[hour] / hourly_avg.mean(), 2),
                    'reason': 'High demand period'
                })
        return {
            'today_energy': round(today_energy, 2),
            'tomorrow_energy': round(hourly_avg.mean() * 24 / 1000, 2),
            'weekly_energy': round(hourly_avg.mean() * 24 * 7 / 1000, 2),
            'peak_hours': peak_hours[:4]
        }
    except Exception as e:
        print(f"Prediction error: {e}")
        return {'today_energy': 0, 'tomorrow_energy': 0, 'weekly_energy': 0, 'peak_hours': []}

def generate_recommendations():
    recommendations = []
    load_power = system_data['load_power']
    battery_soc = system_data['battery_soc']
    battery_voltage = system_data['battery_voltage']
    inverter_power = system_data['inverter_power']
    power_balance = system_data['power_balance']
    current_hour = datetime.now().hour

    if not system_data['esp32_online']:
        recommendations.append({"icon": "\u26a0\ufe0f", "text": "ESP32 is OFFLINE. Check WiFi connection and power supply.", "severity": "critical"})
        return recommendations

    if battery_soc < 15:
        recommendations.append({"icon": "\ud83d\udea8", "text": f"CRITICAL: Battery at {battery_soc:.0f}%. Reduce all non-essential loads immediately.", "severity": "critical"})
    elif battery_soc < 30:
        recommendations.append({"icon": "\u26a0\ufe0f", "text": f"Battery LOW ({battery_soc:.0f}%). Turn off non-essential appliances to conserve power.", "severity": "warning"})
    elif battery_soc < 50:
        recommendations.append({"icon": "\ud83d\udca1", "text": f"Battery at {battery_soc:.0f}%. Moderate level \u2014 use energy-efficient appliances.", "severity": "info"})
    elif battery_soc > 80:
        recommendations.append({"icon": "\u2705", "text": f"Battery well charged ({battery_soc:.0f}%). Good time to run heavy appliances.", "severity": "good"})

    if inverter_power > 0 and load_power > inverter_power:
        recommendations.append({"icon": "\u26a0\ufe0f", "text": f"Power deficit! Load ({load_power:.0f}W) exceeds inverter output ({inverter_power:.0f}W). Battery draining.", "severity": "critical"})
    elif power_balance > 100:
        recommendations.append({"icon": "\u2705", "text": f"Power surplus of {power_balance:.0f}W. Battery is charging.", "severity": "good"})

    if load_power > 700:
        recommendations.append({"icon": "\ud83d\udca1", "text": f"High consumption ({load_power:.0f}W). Consider turning off non-essential loads.", "severity": "warning"})
    elif load_power > 550:
        recommendations.append({"icon": "\ud83d\udca1", "text": f"Moderate consumption ({load_power:.0f}W). Monitor during peak hours.", "severity": "info"})
    elif 0 < load_power < 100:
        recommendations.append({"icon": "\u2705", "text": "Low consumption. Excellent energy practice!", "severity": "good"})

    if 6 <= current_hour < 9:
        recommendations.append({"icon": "\u2600\ufe0f", "text": "Morning hours \u2014 solar panels ramping up. Good time for light appliances."})
    elif 11 <= current_hour <= 14:
        recommendations.append({"icon": "\u2600\ufe0f", "text": "Peak solar production! Run high-power devices now (washing machine, water heater)."})
    elif 17 <= current_hour <= 21:
        recommendations.append({"icon": "\ud83c\udf19", "text": "Evening peak \u2014 solar unavailable. Run heavy appliances before 5PM or after 9PM."})
    elif 22 <= current_hour or current_hour < 5:
        recommendations.append({"icon": "\ud83d\udca4", "text": "Night mode \u2014 minimal solar. Keep loads low to preserve battery until sunrise."})

    if battery_voltage > 0 and battery_voltage < 11.5:
        recommendations.append({"icon": "\ud83d\udd0b", "text": f"Battery voltage low ({battery_voltage:.2f}V). Risk of deep discharge damage.", "severity": "critical"})

    if system_data['weather'].get('daily'):
        today = system_data['weather']['daily'][0]
        if today.get('temp_max', 0) > 35:
            recommendations.append({"icon": "\ud83c\udf21\ufe0f", "text": f"Hot weather ({today['temp_max']}°C). Keep batteries ventilated to prevent overheating."})
        if today.get('uv_index', 0) > 8:
            recommendations.append({"icon": "\u2600\ufe0f", "text": "High UV index \u2014 excellent solar generation expected today."})

    if len(recommendations) < 2:
        recommendations.append({"icon": "\ud83d\udca1", "text": "System operating normally. Schedule heavy appliances during daytime for solar power."})

    return recommendations[:6]

# ========== AI CHATBOT ==========
def chatbot_respond(user_message):
    msg = user_message.lower().strip()
    sd = system_data
    now = datetime.now()

    greetings = ['hello', 'hi', 'hey', 'good morning', 'good evening', 'good afternoon', 'howdy', 'greetings']
    if any(g in msg for g in greetings):
        return f"Hello! I'm your Solar Microgrid AI assistant at Copperbelt University, Kitwe. I can help you with battery status, energy readings, weather, predictions, and energy-saving tips. What would you like to know?"

    if any(w in msg for w in ['battery', 'soc', 'charge', 'charging']):
        soc = sd['battery_soc']
        v = sd['battery_voltage']
        bp = v * sd['battery_current']
        status = "fully charged" if soc > 90 else "well charged" if soc > 60 else "moderate" if soc > 30 else "LOW - charge needed"
        return (f"Battery Status:\n"
                f"\u2022 State of Charge: {soc:.1f}% ({status})\n"
                f"\u2022 Voltage: {v:.2f}V\n"
                f"\u2022 Current: {sd['battery_current']:.2f}A\n"
                f"\u2022 Power: {bp:.1f}W\n"
                f"\u2022 Capacity: 70Ah lead-acid\n"
                f"{'Warning: Battery is low! Reduce non-essential loads.' if soc < 30 else 'Battery is in good condition.'}")

    if any(w in msg for w in ['load', 'consumption', 'power', 'watt', 'how much']):
        return (f"Current Power Readings:\n"
                f"\u2022 Load Power (Consumed): {sd['load_power']:.1f}W\n"
                f"\u2022 Load Current: {sd['load_current']:.2f}A\n"
                f"\u2022 Load Voltage: {sd['load_voltage']:.1f}V\n"
                f"\u2022 Inverter Power (Produced): {sd['inverter_power']:.1f}W\n"
                f"\u2022 Power Balance: {sd['power_balance']:.1f}W\n"
                f"{'Warning: Consumption exceeds production!' if sd['power_balance'] < 0 else 'System is balanced.'}")

    if any(w in msg for w in ['inverter', 'solar', 'panel', 'production', 'produced']):
        return (f"Inverter / Solar Output:\n"
                f"\u2022 Inverter Voltage: {sd['inverter_voltage']:.1f}V\n"
                f"\u2022 Inverter Current: {sd['inverter_current']:.2f}A\n"
                f"\u2022 Inverter Power: {sd['inverter_power']:.1f}W\n"
                f"\u2022 Power Factor: {sd.get('inverter_pf', 0):.2f}\n"
                f"Energy produced so far: {sd['energy_produced_kwh']:.4f}kWh")

    if any(w in msg for w in ['weather', 'temperature', 'temp', 'forecast', 'sun', 'uv']):
        w = sd['weather']
        sl = sd['sunlight']
        daily = w.get('daily', [])
        response = f"Current Weather (Kitwe, Zambia):\n"
        response += f"\u2022 Temperature: {w['current']['temperature']}°C ({w['current']['condition']})\n"
        response += f"\u2022 UV Index: {sl['uv_index']}\n"
        response += f"\u2022 Sunrise: {sl['sunrise']} | Sunset: {sl['sunset']}\n"
        if daily:
            response += f"\u2022 Today: {daily[0]['temp_min']}°C - {daily[0]['temp_max']}°C\n"
            response += f"{'High solar production expected today!' if daily[0].get('uv_index', 0) > 6 else 'Moderate solar expected.'}"
        return response

    if any(w in msg for w in ['predict', 'prediction', 'forecast energy', 'tomorrow', 'weekly', 'today energy']):
        p = sd['predictions']
        return (f"Energy Predictions:\n"
                f"\u2022 Today's predicted: {p.get('today_energy', 0)}kWh\n"
                f"\u2022 Tomorrow's predicted: {p.get('tomorrow_energy', 0)}kWh\n"
                f"\u2022 Weekly predicted: {p.get('weekly_energy', 0)}kWh\n"
                f"\u2022 Peak hours: {', '.join([h['hour'] for h in p.get('peak_hours', [])]) or 'N/A'}\n"
                f"Tip: Run high-power devices during peak solar hours (11AM-2PM) for best results.")

    if any(w in msg for w in ['relay', 'switch', 'turn on', 'turn off', 'control']):
        return (f"Relay Status:\n"
                f"\u2022 Relay 1 (Essential): {sd['load1_state']}\n"
                f"\u2022 Relay 2 (Priority): {sd['load2_state']}\n"
                f"To control relays, use the Relay Control buttons on the dashboard.\n"
                f"Relay 1 powers essential loads (always on preferred).\n"
                f"Relay 2 powers priority loads (can be switched off to save power).")

    if any(w in msg for w in ['save', 'tip', 'advice', 'reduce', 'efficiency', 'efficient']):
        tips = [
            "1. Run heavy appliances (washing machine, water heater) between 11AM-2PM when solar production peaks.",
            "2. Switch off lights and fans when not in use \u2014 even small loads add up at night.",
            "3. Use LED bulbs instead of incandescent \u2014 they use 80% less power.",
            "4. Keep battery charged above 30% to maximize battery lifespan.",
            "5. Avoid running all appliances simultaneously \u2014 stagger usage to prevent overload.",
            "6. Monitor the power balance \u2014 if negative, battery is draining. Reduce load.",
            "7. In hot weather, keep batteries in a ventilated area to prevent thermal damage.",
        ]
        return "Energy Saving Tips:\n" + "\n".join(tips)

    if any(w in msg for w in ['status', 'system', 'overview', 'summary', 'how is']):
        online = "ONLINE" if sd['esp32_online'] else "OFFLINE"
        return (f"System Overview:\n"
                f"\u2022 ESP32 Controller: {online}\n"
                f"\u2022 Battery: {sd['battery_soc']:.1f}% ({sd['battery_voltage']:.2f}V)\n"
                f"\u2022 Load: {sd['load_power']:.1f}W | Inverter: {sd['inverter_power']:.1f}W\n"
                f"\u2022 Relay 1: {sd['load1_state']} | Relay 2: {sd['load2_state']}\n"
                f"\u2022 Power Balance: {sd['power_balance']:.1f}W\n"
                f"\u2022 Last Update: {sd['timestamp']}\n"
                f"Everything {'looks good!' if sd['esp32_online'] and sd['battery_soc'] > 30 else 'needs attention.'}")

    if any(w in msg for w in ['cost', 'saving', 'zkw', 'money', 'bill']):
        energy = sd['energy_produced_kwh']
        cost = energy * ZMW_PER_KWH
        co2 = energy * KG_CO2_PER_KWH
        return (f"Energy Savings (Copperbelt):\n"
                f"\u2022 Energy Produced: {energy:.4f}kWh\n"
                f"\u2022 Cost Savings: ZMW {cost:.2f} (@ ZMW {ZMW_PER_KWH}/kWh)\n"
                f"\u2022 CO2 Offset: {co2:.2f}kg\n"
                f"Your solar microgrid reduces electricity costs and carbon emissions!")

    if any(w in msg for w in ['help', 'what can you', 'command', 'ask']):
        return ("I can help with:\n"
                "\u2022 Battery status \u2014 'What's my battery level?'\n"
                "\u2022 Power readings \u2014 'How much power am I using?'\n"
                "\u2022 Solar output \u2014 'What's the solar production?'\n"
                "\u2022 Weather \u2014 'What's the weather forecast?'\n"
                "\u2022 Predictions \u2014 'What are the energy predictions?'\n"
                "\u2022 Relay control \u2014 'What are the relay statuses?'\n"
                "\u2022 Energy tips \u2014 'Give me energy saving tips'\n"
                "\u2022 Cost savings \u2014 'How much money am I saving?'\n"
                "\u2022 System overview \u2014 'Give me a system summary'")

    if any(w in msg for w in ['thank', 'thanks', 'appreciate']):
        return "You're welcome! I'm here to help you manage your solar microgrid efficiently. Feel free to ask anything!"

    if any(w in msg for w in ['time', 'date', 'when', 'now']):
        return f"Current time: {now.strftime('%Y-%m-%d %H:%M:%S')} (Africa/Lusaka timezone)"

    return (f"I'm not sure I understand that. You can ask me about:\n"
            f"\u2022 Battery status and charge level\n"
            f"\u2022 Power consumption and production\n"
            f"\u2022 Weather and solar forecast\n"
            f"\u2022 Energy predictions\n"
            f"\u2022 Relay control status\n"
            f"\u2022 Energy saving tips\n"
            f"\u2022 Cost savings\n\n"
            f"Type 'help' for a full list of commands.")

@app.route('/api/chat', methods=['POST'])
def chat_endpoint():
    try:
        data = request.get_json(force=True)
        user_msg = data.get('message', '').strip()
        if not user_msg:
            return jsonify({'error': 'Empty message'}), 400
        response = chatbot_respond(user_msg)
        return jsonify({'response': response, 'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')})
    except Exception as e:
        return jsonify({'response': f'Error: {str(e)}'}), 500

# ========== WEATHER DATA ==========
def get_weather_data():
    url = "https://api.open-meteo.com/v1/forecast"
    params = {
        "latitude": COPPERBELT_LAT, "longitude": COPPERBELT_LON,
        "daily": ["temperature_2m_max", "temperature_2m_min", "sunrise", "sunset", "uv_index_max"],
        "timezone": "Africa/Lusaka", "forecast_days": 7
    }
    try:
        response = requests.get(url, params=params, timeout=10)
        if response.status_code == 200:
            data = response.json()
            daily = data['daily']
            system_data['weather']['daily'] = []
            for i in range(len(daily['time'])):
                system_data['weather']['daily'].append({
                    'date': daily['time'][i],
                    'temp_max': daily['temperature_2m_max'][i],
                    'temp_min': daily['temperature_2m_min'][i],
                    'sunrise': daily['sunrise'][i],
                    'sunset': daily['sunset'][i],
                    'uv_index': daily['uv_index_max'][i]
                })
            system_data['weather']['current']['temperature'] = daily['temperature_2m_max'][0]
            if daily['temperature_2m_max'][0] > 30:
                system_data['weather']['current']['condition'] = "Hot"
            elif daily['temperature_2m_max'][0] < 15:
                system_data['weather']['current']['condition'] = "Cool"
            else:
                system_data['weather']['current']['condition'] = "Mild"
            system_data['sunlight'] = {
                'uv_index': daily['uv_index_max'][0],
                'sunrise': daily['sunrise'][0].split('T')[1][:5],
                'sunset': daily['sunset'][0].split('T')[1][:5]
            }
    except Exception as e:
        print(f"Weather API error: {e}")

def update_weather_periodically():
    while True:
        get_weather_data()
        time.sleep(3600)

def check_esp32_status():
    global email_alert_state
    while True:
        time.sleep(15)
        try:
            if system_data['last_esp32_seen']:
                last = datetime.strptime(system_data['last_esp32_seen'], '%Y-%m-%d %H:%M:%S')
                if (datetime.now() - last).total_seconds() > 20:
                    if system_data['esp32_online']:
                        system_data['esp32_online'] = False
                        system_data['recommendations'] = generate_recommendations()
                        print("ESP32 went OFFLINE")
                        if not email_alert_state['esp32_was_online']:
                            pass
                        else:
                            notify_esp32_offline()
                            email_alert_state['esp32_was_online'] = False
                else:
                    if not email_alert_state['esp32_was_online'] and system_data['esp32_online']:
                        email_alert_state['esp32_was_online'] = True
                        notify_esp32_online()
                        email_alert_state['low_battery_sent'] = False
            elif system_data['esp32_online']:
                system_data['esp32_online'] = False
        except Exception as e:
            print(f"ESP32 status check error: {e}")

def check_battery_email():
    while True:
        time.sleep(60)
        try:
            soc = system_data['battery_soc']
            if system_data['esp32_online'] and soc > 0 and soc < 20 and not email_alert_state['low_battery_sent']:
                notify_low_battery(soc)
                email_alert_state['low_battery_sent'] = True
            elif soc >= 25:
                email_alert_state['low_battery_sent'] = False
        except Exception as e:
            print(f"Battery email check error: {e}")

# ========== FLASK API ROUTES ==========
@app.route('/')
def dashboard():
    return render_template_string(HTML_TEMPLATE)

@app.route('/api/ping', methods=['GET', 'POST'])
def ping():
    return jsonify({'status': 'ok', 'time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')})

@app.route('/api/update', methods=['GET'])
def update_from_esp32():
    args = request.args
    if not args or len(args) < 3:
        return jsonify({'error': 'No query params'}), 400
    try:
        system_data['inverter_voltage'] = float(args.get('iv', 0))
        system_data['inverter_current'] = float(args.get('ic', 0))
        system_data['inverter_power'] = float(args.get('ip', 0))
        system_data['energy_produced_kwh'] = float(args.get('ie', 0))
        system_data['load_voltage'] = float(args.get('lv', 0))
        system_data['load_current'] = float(args.get('lc', 0))
        system_data['load_power'] = float(args.get('lp', 0))
        system_data['energy_consumed_kwh'] = float(args.get('le', 0))
        system_data['battery_voltage'] = float(args.get('bv', 0))
        system_data['battery_current'] = float(args.get('bc', 0))
        system_data['battery_soc'] = float(args.get('bs', 0))
        system_data['load1_state'] = args.get('r1', 'OFF')
        system_data['load2_state'] = args.get('r2', 'OFF')
        system_data['power_balance'] = system_data['inverter_power'] - system_data['load_power']
        system_data['load1_power'] = system_data['load_power']
        system_data['load2_power'] = 0
        system_data['trip_state'] = 'NOR'
        system_data['phase'] = 'P1'
        system_data['esp32_online'] = True
        system_data['last_esp32_seen'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        system_data['timestamp'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        real_time_power.append({
            'timestamp': datetime.now(),
            'consumed': system_data['load_power'],
            'produced': system_data['inverter_power']
        })

        log_sensor_reading()
        system_data['predictions'] = calculate_energy_predictions()
        system_data['recommendations'] = generate_recommendations()

        return jsonify({'status': 'ok'}), 200
    except Exception as e:
        print(f"Update error: {e}")
        return jsonify({'error': str(e)}), 400

@app.route('/api/data', methods=['GET', 'POST'])
def handle_data():
    if request.method == 'POST':
        try:
            raw = request.get_data(as_text=True)
            data = json.loads(raw)
        except Exception as e:
            return jsonify({'error': str(e)}), 400
        if not data or len(data) < 3:
            return jsonify({'error': 'Empty data'}), 400
        system_data['battery_voltage'] = float(data.get('battery_voltage', 0))
        system_data['battery_current'] = float(data.get('battery_current', 0))
        system_data['battery_soc'] = float(data.get('battery_soc', 0))
        system_data['inverter_voltage'] = float(data.get('inverter_voltage', 0))
        system_data['inverter_current'] = float(data.get('inverter_current', 0))
        system_data['inverter_power'] = float(data.get('inverter_power', 0))
        system_data['load_voltage'] = float(data.get('load_voltage', 0))
        system_data['load_current'] = float(data.get('load_current', 0))
        system_data['load_power'] = float(data.get('load_power', 0))
        system_data['energy_consumed_kwh'] = float(data.get('load_energy', 0))
        system_data['energy_produced_kwh'] = float(data.get('inverter_energy', 0))
        system_data['load1_state'] = data.get('relay1_state', 'OFF')
        system_data['load2_state'] = data.get('relay2_state', 'OFF')
        system_data['power_balance'] = system_data['inverter_power'] - system_data['load_power']
        system_data['esp32_online'] = True
        system_data['last_esp32_seen'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        system_data['timestamp'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        log_sensor_reading()
        system_data['predictions'] = calculate_energy_predictions()
        system_data['recommendations'] = generate_recommendations()
        return jsonify({'status': 'ok'}), 200
    return jsonify(system_data)

@app.route('/api/commands', methods=['GET'])
def get_commands():
    with command_lock:
        if command_queue:
            cmd = command_queue.pop(0)
            return jsonify(cmd)
        return jsonify({'cmd': 'none'})

@app.route('/api/relay', methods=['POST'])
def control_relay():
    data = request.get_json(force=True)
    if not data:
        return jsonify({'error': 'No data'}), 400
    relay = data.get('relay')
    state = data.get('state', 'ON')
    if relay not in [1, 2]:
        return jsonify({'error': 'Invalid relay'}), 400
    if state not in ['ON', 'OFF']:
        return jsonify({'error': 'Invalid state'}), 400
    cmd = {'cmd': 'relay', 'relay': relay, 'state': state, 'id': int(time.time() * 1000)}
    with command_lock:
        command_queue.append(cmd)
    log_relay_event(relay, state, 'manual')
    print(f"Relay command: Relay {relay} -> {state}")
    return jsonify({'status': 'ok', 'cmd': cmd})

@app.route('/api/relay/ack', methods=['GET', 'POST'])
def relay_ack():
    if request.method == 'GET':
        relay = request.args.get('relay', type=int)
        state = request.args.get('state', 'OFF')
    else:
        data = request.get_json(force=True)
        if not data:
            return jsonify({'error': 'No data'}), 400
        relay = data.get('relay')
        state = data.get('state', 'OFF')
    if relay == 1:
        system_data['load1_state'] = state
    elif relay == 2:
        system_data['load2_state'] = state
    print(f"Relay ACK: Relay {relay} -> {state}")
    return jsonify({'status': 'ok'})

@app.route('/api/real_time_power')
def get_real_time_power():
    consumed = [p['consumed'] for p in real_time_power]
    produced = [p['produced'] for p in real_time_power]
    timestamps = [p['timestamp'].strftime('%H:%M:%S') for p in real_time_power]
    return jsonify({'consumed': consumed, 'produced': produced, 'timestamps': timestamps})

@app.route('/api/historical_data')
def get_historical_data():
    if len(historical_power) > 0:
        df = pd.DataFrame(historical_power)
        df['date'] = df['timestamp'].dt.date
        daily_avg = df.groupby('date').agg({'consumed': 'mean', 'produced': 'mean'}).reset_index()
        return jsonify({
            'timestamps': daily_avg['date'].astype(str).tolist(),
            'consumed': daily_avg['consumed'].tolist(),
            'produced': daily_avg['produced'].tolist()
        })
    return jsonify({'timestamps': [], 'consumed': [], 'produced': []})

@app.route('/api/comparison')
def get_comparison():
    try:
        now = datetime.now()
        this_month_start = now.replace(day=1, hour=0, minute=0, second=0)
        last_month_end = this_month_start - timedelta(seconds=1)
        last_month_start = last_month_end.replace(day=1, hour=0, minute=0, second=0)

        this_rows = get_readings_between(this_month_start.strftime('%Y-%m-%d'), now.strftime('%Y-%m-%d %H:%M:%S'))
        last_rows = get_readings_between(last_month_start.strftime('%Y-%m-%d'), last_month_end.strftime('%Y-%m-%d %H:%M:%S'))

        def daily_avg(rows):
            if not rows:
                return {'dates': [], 'consumed': [], 'produced': []}
            df = pd.DataFrame(rows)
            df['date'] = pd.to_datetime(df['timestamp']).dt.date
            daily = df.groupby('date').agg({'load_power': 'mean', 'inverter_power': 'mean'}).reset_index()
            return {
                'dates': daily['date'].astype(str).tolist(),
                'consumed': daily['load_power'].round(1).tolist(),
                'produced': daily['inverter_power'].round(1).tolist()
            }

        return jsonify({
            'this_month': daily_avg(this_rows),
            'last_month': daily_avg(last_rows),
            'this_month_label': this_month_start.strftime('%B %Y'),
            'last_month_label': last_month_start.strftime('%B %Y')
        })
    except Exception as e:
        print(f"Comparison error: {e}")
        return jsonify({
            'this_month': {'dates': [], 'consumed': [], 'produced': []},
            'last_month': {'dates': [], 'consumed': [], 'produced': []},
            'this_month_label': datetime.now().strftime('%B %Y'),
            'last_month_label': (datetime.now() - timedelta(days=30)).strftime('%B %Y')
        })

@app.route('/api/export/csv')
def export_csv():
    try:
        start = request.args.get('start', (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
        end = request.args.get('end', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        rows = get_readings_between(start, end)
        output = BytesIO()
        writer_text = ['timestamp,battery_voltage,battery_current,battery_soc,inverter_voltage,inverter_current,inverter_power,load_voltage,load_current,load_power,energy_consumed_kwh,energy_produced_kwh,relay1_state,relay2_state,power_balance\n']
        for r in rows:
            writer_text.append(f"{r['timestamp']},{r['battery_voltage']},{r['battery_current']},{r['battery_soc']},{r['inverter_voltage']},{r['inverter_current']},{r['inverter_power']},{r['load_voltage']},{r['load_current']},{r['load_power']},{r['energy_consumed_kwh']},{r['energy_produced_kwh']},{r['relay1_state']},{r['relay2_state']},{r['power_balance']}\n")
        output.write(''.join(writer_text).encode('utf-8'))
        output.seek(0)
        filename = f"solar_data_{start}_{end[:10]}.csv"
        return send_file(output, mimetype='text/csv', as_attachment=True, download_name=filename)
    except Exception as e:
        print(f"CSV export error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/export/excel')
def export_excel():
    try:
        start = request.args.get('start', (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
        end = request.args.get('end', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        rows = get_readings_between(start, end)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sensor Readings"
        headers = ['Timestamp', 'Battery V', 'Battery A', 'SOC%', 'Inv V', 'Inv A', 'Inv W', 'Load V', 'Load A', 'Load W', 'Cons kWh', 'Prod kWh', 'R1', 'R2', 'Balance W']
        header_fill = PatternFill(start_color='1a1a2e', end_color='1a1a2e', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        for i, r in enumerate(rows, 2):
            ws.cell(row=i, column=1, value=r['timestamp'])
            ws.cell(row=i, column=2, value=r['battery_voltage'])
            ws.cell(row=i, column=3, value=r['battery_current'])
            ws.cell(row=i, column=4, value=r['battery_soc'])
            ws.cell(row=i, column=5, value=r['inverter_voltage'])
            ws.cell(row=i, column=6, value=r['inverter_current'])
            ws.cell(row=i, column=7, value=r['inverter_power'])
            ws.cell(row=i, column=8, value=r['load_voltage'])
            ws.cell(row=i, column=9, value=r['load_current'])
            ws.cell(row=i, column=10, value=r['load_power'])
            ws.cell(row=i, column=11, value=r['energy_consumed_kwh'])
            ws.cell(row=i, column=12, value=r['energy_produced_kwh'])
            ws.cell(row=i, column=13, value=r['relay1_state'])
            ws.cell(row=i, column=14, value=r['relay2_state'])
            ws.cell(row=i, column=15, value=r['power_balance'])

        if rows:
            df = pd.DataFrame(rows)
            ws2 = wb.create_sheet("Summary")
            ws2.cell(row=1, column=1, value="Metric").fill = header_fill
            ws2.cell(row=1, column=1).font = header_font
            ws2.cell(row=1, column=2, value="Average").fill = header_fill
            ws2.cell(row=1, column=2).font = header_font
            ws2.cell(row=1, column=3, value="Min").fill = header_fill
            ws2.cell(row=1, column=3).font = header_font
            ws2.cell(row=1, column=4, value="Max").fill = header_fill
            ws2.cell(row=1, column=4).font = header_font
            metrics = [('Battery SOC%', 'battery_soc'), ('Inverter Power W', 'inverter_power'),
                       ('Load Power W', 'load_power'), ('Battery Voltage V', 'battery_voltage')]
            for i, (name, col_name) in enumerate(metrics, 2):
                ws2.cell(row=i, column=1, value=name)
                ws2.cell(row=i, column=2, value=round(df[col_name].mean(), 2))
                ws2.cell(row=i, column=3, value=round(df[col_name].min(), 2))
                ws2.cell(row=i, column=4, value=round(df[col_name].max(), 2))

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"solar_data_{start}_{end[:10]}.xlsx"
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=filename)
    except Exception as e:
        print(f"Excel export error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/report/monthly')
def monthly_report():
    try:
        month = request.args.get('month', datetime.now().month, type=int)
        year = request.args.get('year', datetime.now().year, type=int)
        start_date = f"{year}-{month:02d}-01"
        last_day = calendar.monthrange(year, month)[1]
        end_date = f"{year}-{month:02d}-{last_day:02d} 23:59:59"
        rows = get_readings_between(start_date, end_date)
        relay_events = get_relay_events_between(start_date, end_date)

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=20*mm, bottomMargin=20*mm)
        styles = getSampleStyleSheet()
        elements = []

        title_style = ParagraphStyle('Title2', parent=styles['Title'], fontSize=20, spaceAfter=10)
        subtitle_style = ParagraphStyle('Subtitle2', parent=styles['Normal'], fontSize=12, alignment=TA_CENTER, spaceAfter=20, textColor=colors.grey)
        elements.append(Paragraph("Copperbelt University - Solar Microgrid", title_style))
        elements.append(Paragraph(f"Monthly Energy Report - {calendar.month_name[month]} {year}", subtitle_style))
        elements.append(Paragraph(f"Location: Kitwe, Zambia | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", subtitle_style))
        elements.append(Spacer(1, 10*mm))

        elements.append(Paragraph("Summary Statistics", styles['Heading2']))
        if rows:
            df = pd.DataFrame(rows)
            summary_data = [
                ['Metric', 'Average', 'Min', 'Max', 'Unit'],
                ['Battery SOC', f"{df['battery_soc'].mean():.1f}", f"{df['battery_soc'].min():.1f}", f"{df['battery_soc'].max():.1f}", '%'],
                ['Battery Voltage', f"{df['battery_voltage'].mean():.2f}", f"{df['battery_voltage'].min():.2f}", f"{df['battery_voltage'].max():.2f}", 'V'],
                ['Inverter Power', f"{df['inverter_power'].mean():.1f}", f"{df['inverter_power'].min():.1f}", f"{df['inverter_power'].max():.1f}", 'W'],
                ['Load Power', f"{df['load_power'].mean():.1f}", f"{df['load_power'].min():.1f}", f"{df['load_power'].max():.1f}", 'W'],
                ['Power Balance', f"{df['power_balance'].mean():.1f}", f"{df['power_balance'].min():.1f}", f"{df['power_balance'].max():.1f}", 'W'],
            ]
            total_energy_consumed = df['load_power'].mean() * len(rows) * 5 / 3600 / 1000
            total_energy_produced = df['inverter_power'].mean() * len(rows) * 5 / 3600 / 1000
            cost_savings = total_energy_produced * ZMW_PER_KWH
            co2_offset = total_energy_produced * KG_CO2_PER_KWH

            summary_data.append(['Total Energy Consumed', f"{total_energy_consumed:.2f}", '', '', 'kWh'])
            summary_data.append(['Total Energy Produced', f"{total_energy_produced:.2f}", '', '', 'kWh'])
            summary_data.append(['Cost Savings', f"{cost_savings:.2f}", '', '', 'ZMW'])
            summary_data.append(['CO2 Offset', f"{co2_offset:.2f}", '', '', 'kg'])

            table = Table(summary_data, colWidths=[120, 70, 70, 70, 50])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1a2e')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
            ]))
            elements.append(table)
        else:
            elements.append(Paragraph("No data available for this period.", styles['Normal']))

        elements.append(Spacer(1, 10*mm))
        elements.append(Paragraph("Relay Events", styles['Heading2']))
        if relay_events:
            relay_data = [['Timestamp', 'Relay', 'State', 'Source']]
            for e in relay_events[:50]:
                relay_data.append([e['timestamp'], str(e['relay']), e['state'], e['source']])
            relay_table = Table(relay_data, colWidths=[130, 50, 50, 100])
            relay_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2196F3')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
            ]))
            elements.append(relay_table)
        else:
            elements.append(Paragraph("No relay events this month.", styles['Normal']))

        elements.append(Spacer(1, 15*mm))
        elements.append(Paragraph(f"Tariff: ZMW {ZMW_PER_KWH}/kWh | CO2 Factor: {KG_CO2_PER_KWH} kg/kWh", styles['Normal']))

        doc.build(elements)
        buffer.seek(0)
        return send_file(buffer, mimetype='application/pdf', as_attachment=True,
                         download_name=f"solar_report_{year}_{month:02d}.pdf")
    except Exception as e:
        print(f"PDF report error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/download_csv')
def download_csv():
    if os.path.exists(HISTORICAL_DATA_FILE):
        return send_file(HISTORICAL_DATA_FILE, as_attachment=True)
    return "No data available", 404

@app.route('/api/test_email', methods=['POST'])
def test_email():
    try:
        result = send_email("Solar Microgrid: Test Email",
            '<div style="font-family:Arial;padding:20px"><h2>Test Email</h2><p>Email notifications are working correctly.</p><p>Time: ' + datetime.now().strftime('%Y-%m-%d %H:%M:%S') + '</p></div>')
        if result:
            return jsonify({'status': 'ok', 'message': 'Email sent successfully'})
        return jsonify({'status': 'error', 'message': 'Email send failed - check logs'}), 500
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

# ========== HTML DASHBOARD ==========
HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Solar Microgrid - Smart Energy Management</title>
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
    <script src="https://cdn.plot.ly/plotly-latest.min.js"></script>
    <style>
        *{margin:0;padding:0;box-sizing:border-box}
        body{font-family:'Segoe UI',Tahoma,sans-serif;background:linear-gradient(135deg,#1a1a2e,#16213e);min-height:100vh;padding:20px}
        .container{max-width:1600px;margin:0 auto}
        header{text-align:center;color:white;margin-bottom:30px}
        h1{font-size:2.5em;margin-bottom:10px}
        .subtitle{font-size:1.1em;opacity:.9}
        .grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(380px,1fr));gap:25px;margin-bottom:25px}
        .card{background:white;border-radius:20px;padding:25px;box-shadow:0 10px 40px rgba(0,0,0,.2);transition:transform .3s}
        .card:hover{transform:translateY(-5px)}
        .card-header{font-size:1.3em;font-weight:bold;color:#1a1a2e;margin-bottom:20px;border-bottom:3px solid #1a1a2e;padding-bottom:12px}
        .value{font-size:2.2em;font-weight:bold;color:#333;margin:10px 0}
        .unit{font-size:.5em;color:#666}
        .status{padding:8px 15px;border-radius:8px;display:inline-block;font-weight:bold}
        .status-normal{background:#4CAF50;color:white}
        .status-overload{background:#ff9800;color:white}
        .status-tripped{background:#f44336;color:white;animation:blink 1s infinite}
        @keyframes blink{0%,100%{opacity:1}50%{opacity:.5}}
        .prediction-card{background:linear-gradient(135deg,#667eea,#764ba2);color:white;padding:15px;border-radius:12px;margin:10px 0}
        .chart-container{margin:20px 0;height:500px}
        .recommendation{background:#e3f2fd;padding:12px;border-radius:10px;margin:10px 0;border-left:4px solid #2196F3}
        .peak-hour{display:inline-block;background:#ff9800;color:white;padding:6px 12px;border-radius:8px;margin:5px;font-size:13px}
        .load-badge{display:inline-block;padding:4px 10px;border-radius:20px;margin:3px;font-size:12px;font-weight:bold}
        .load-on{background:#4CAF50;color:white}
        .load-off{background:#9e9e9e;color:white}
        .relay-btn{display:inline-block;padding:8px 20px;border:none;border-radius:8px;font-size:14px;font-weight:bold;cursor:pointer;margin:5px;transition:all .2s}
        .relay-btn-on{background:#4CAF50;color:white}
        .relay-btn-on:hover{background:#388E3C}
        .relay-btn-off{background:#f44336;color:white}
        .relay-btn-off:hover{background:#d32f2f}
        .relay-btn:disabled{background:#ccc;cursor:not-allowed}
        .timestamp{font-size:.8em;color:#666;margin-top:20px;text-align:center}
        footer{text-align:center;color:white;margin-top:30px;padding:20px}
        .balance-positive{color:#4CAF50}
        .balance-negative{color:#f44336}
        .esp-status{margin-top:10px;font-size:13px}
        .export-btn{display:inline-block;padding:10px 20px;border:none;border-radius:8px;font-size:14px;font-weight:bold;cursor:pointer;margin:5px;color:white;transition:all .2s}
        .export-csv{background:#2196F3}.export-csv:hover{background:#1976D2}
        .export-excel{background:#4CAF50}.export-excel:hover{background:#388E3C}
        .export-pdf{background:#f44336}.export-pdf:hover{background:#d32f2f}
        .config-input{width:60px;padding:5px;border:1px solid #ccc;border-radius:4px;margin:5px}
        .config-label{font-size:13px;color:#666}
        .email-btn{background:#9c27b0;color:white;padding:10px 20px;border:none;border-radius:8px;font-size:14px;font-weight:bold;cursor:pointer;margin:5px;transition:all .2s}
        .email-btn:hover{background:#7b1fa2}
        .email-btn:disabled{background:#ccc;cursor:not-allowed}
        .chat-card{position:relative}
        .chat-messages{height:350px;overflow-y:auto;padding:10px;background:#f5f5f5;border-radius:12px;margin-bottom:10px;scroll-behavior:smooth}
        .chat-msg{margin:8px 0;display:flex;gap:8px}
        .chat-msg.user{justify-content:flex-end}
        .chat-bubble{max-width:80%;padding:10px 14px;border-radius:16px;font-size:13px;line-height:1.5;white-space:pre-line;word-wrap:break-word}
        .chat-bubble.bot{background:#1a1a2e;color:white;border-bottom-left-radius:4px}
        .chat-bubble.user{background:#2196F3;color:white;border-bottom-right-radius:4px}
        .chat-avatar{width:30px;height:30px;border-radius:50%;display:flex;align-items:center;justify-content:center;font-size:14px;flex-shrink:0}
        .chat-avatar.bot-avatar{background:#1a1a2e;color:white}
        .chat-avatar.user-avatar{background:#2196F3;color:white}
        .chat-input-row{display:flex;gap:8px}
        .chat-input{flex:1;padding:10px 14px;border:2px solid #ddd;border-radius:25px;font-size:14px;outline:none;transition:border .2s}
        .chat-input:focus{border-color:#2196F3}
        .chat-send{width:42px;height:42px;border-radius:50%;border:none;background:#2196F3;color:white;font-size:18px;cursor:pointer;transition:background .2s;display:flex;align-items:center;justify-content:center}
        .chat-send:hover{background:#1976D2}
        .chat-send:disabled{background:#ccc;cursor:not-allowed}
        .chat-suggestions{display:flex;flex-wrap:wrap;gap:5px;margin-bottom:8px}
        .chat-suggest{padding:4px 10px;border:1px solid #ddd;border-radius:15px;font-size:11px;cursor:pointer;background:white;transition:all .2s;color:#333}
        .chat-suggest:hover{background:#2196F3;color:white;border-color:#2196F3}
        .rec-icon{margin-right:8px;font-size:16px}
        .rec-critical{background:#fff3f3;border-left-color:#f44336}
        .rec-warning{background:#fff8e1;border-left-color:#ff9800}
        .rec-good{background:#f1f8e9;border-left-color:#4CAF50}
        .rec-info{background:#e3f2fd;border-left-color:#2196F3}
    </style>
</head>
<body>
    <div class="container">
        <header>
            <h1>Solar Powered Microgrid</h1>
            <div class="subtitle">Copperbelt University, Kitwe | Real-time Monitoring & Energy Predictions</div>
        </header>

        <div class="grid">
            <div class="card">
                <div class="card-header">Battery Status</div>
                <div id="esp-status" class="esp-status"></div>
                <div class="value" id="battery-soc">0<span class="unit">%</span></div>
                <div>State of Charge</div>
                <div class="value" id="battery-voltage">0<span class="unit">V</span></div>
                <div>Battery Voltage</div>
                <div class="value" id="battery-current">0<span class="unit">A</span></div>
                <div>Battery Current</div>
            </div>
            <div class="card">
                <div class="card-header">Power Dashboard</div>
                <div class="value" id="load-power">0<span class="unit">W</span></div>
                <div>Load Power (Consumed)</div>
                <div class="value" id="inverter-power">0<span class="unit">W</span></div>
                <div>Inverter Power (Produced)</div>
                <div class="value" id="power-balance">0<span class="unit">W</span></div>
                <div>Power Balance</div>
                <div id="trip-status" class="status status-normal" style="margin-top:10px">NORMAL</div>
            </div>
            <div class="card">
                <div class="card-header">Load Status & Control</div>
                <div class="value" id="load-current">0<span class="unit">A</span></div>
                <div>Total Load Current</div>
                <div style="margin:15px 0">
                    <span class="load-badge load-off" id="load1-badge">Load 1: OFF</span>
                    <span class="load-badge load-off" id="load2-badge">Load 2: OFF</span>
                </div>
                <div id="load-powers" style="font-size:14px;color:#666"></div>
                <div class="card-header" style="margin-top:15px;border-bottom:2px solid #2196F3">Relay Control</div>
                <div style="margin:10px 0">
                    <div style="margin-bottom:10px">
                        <strong>Relay 1 (Essential):</strong>
                        <button class="relay-btn relay-btn-on" onclick="sendRelayCmd(1,'ON')">ON</button>
                        <button class="relay-btn relay-btn-off" onclick="sendRelayCmd(1,'OFF')">OFF</button>
                    </div>
                    <div>
                        <strong>Relay 2 (Priority):</strong>
                        <button class="relay-btn relay-btn-on" onclick="sendRelayCmd(2,'ON')">ON</button>
                        <button class="relay-btn relay-btn-off" onclick="sendRelayCmd(2,'OFF')">OFF</button>
                    </div>
                </div>
                <div class="value" id="energy-consumed">0.000000<span class="unit">kWh</span></div>
                <div>Total Energy Consumed</div>
                <div class="value" id="energy-produced">0.000000<span class="unit">kWh</span></div>
                <div>Total Energy Produced</div>
            </div>
        </div>

        <div class="grid">
            <div class="card">
                <div class="card-header">Export & Reports</div>
                <div style="margin:10px 0">
                    <div class="config-label" style="margin-bottom:5px">Date Range:</div>
                    <input type="date" id="export-start" style="padding:5px;border:1px solid #ccc;border-radius:4px">
                    <span style="margin:0 5px">to</span>
                    <input type="date" id="export-end" style="padding:5px;border:1px solid #ccc;border-radius:4px">
                </div>
                <div style="margin-top:15px">
                    <button class="export-btn export-csv" onclick="exportCSV()">Export CSV</button>
                    <button class="export-btn export-excel" onclick="exportExcel()">Export Excel</button>
                </div>
                <div style="margin-top:15px">
                    <div class="config-label" style="margin-bottom:5px">Monthly PDF Report:</div>
                    <select id="report-month" style="padding:5px;border:1px solid #ccc;border-radius:4px">
                        <option value="1">January</option><option value="2">February</option><option value="3">March</option>
                        <option value="4">April</option><option value="5">May</option><option value="6">June</option>
                        <option value="7">July</option><option value="8">August</option><option value="9">September</option>
                        <option value="10">October</option><option value="11">November</option><option value="12">December</option>
                    </select>
                    <input type="number" id="report-year" value="2026" style="width:70px;padding:5px;border:1px solid #ccc;border-radius:4px">
                    <button class="export-btn export-pdf" onclick="exportPDF()">Download PDF</button>
                </div>
            </div>
            <div class="card">
                <div class="card-header">Energy Predictions</div>
                <div class="prediction-card"><div>Today's Predicted Energy</div><div class="value" id="pred-today">0<span class="unit">kWh</span></div></div>
                <div class="prediction-card"><div>Tomorrow's Predicted Energy</div><div class="value" id="pred-tomorrow">0<span class="unit">kWh</span></div></div>
                <div class="prediction-card"><div>Weekly Predicted Energy</div><div class="value" id="pred-weekly">0<span class="unit">kWh</span></div></div>
                <div class="card-header" style="margin-top:15px">Peak Usage Hours</div>
                <div id="peak-hours"></div>
            </div>
            <div class="card">
                <div class="card-header">Notifications</div>
                <p style="color:#666;font-size:13px;margin-bottom:15px">Email alerts are sent automatically when ESP32 goes offline or battery is low.</p>
                <button class="email-btn" onclick="testEmail()" id="email-test-btn">Send Test Email</button>
                <div id="email-status" style="margin-top:10px;font-size:13px"></div>
            </div>
        </div>

        <div class="grid">
            <div class="card">
                <div class="card-header">Real-time Power Comparison</div>
                <div class="chart-container"><canvas id="powerChart"></canvas></div>
            </div>
            <div class="card">
                <div class="card-header">Monthly Comparison (This vs Last)</div>
                <div id="comparisonChart" style="height:500px"></div>
            </div>
        </div>

        <div class="grid">
            <div class="card">
                <div class="card-header">Historical Power Trends</div>
                <div id="historicalChart" style="height:450px"></div>
            </div>
            <div class="card">
                <div class="card-header">Energy Recommendations</div>
                <div id="recommendations"></div>
                <div class="card-header" style="margin-top:15px">Weather</div>
                <div id="weather-info"></div>
                <div id="sunlight-info"></div>
            </div>
            <div class="card chat-card">
                <div class="card-header">Solar AI Assistant</div>
                <div class="chat-messages" id="chat-messages">
                    <div class="chat-msg">
                        <div class="chat-avatar bot-avatar">AI</div>
                        <div class="chat-bubble bot">Hello! I'm your Solar Microgrid AI assistant. Ask me about battery status, power readings, weather, energy tips, or system status.</div>
                    </div>
                </div>
                <div class="chat-suggestions" id="chat-suggestions">
                    <span class="chat-suggest" onclick="sendChat('battery status')">Battery Status</span>
                    <span class="chat-suggest" onclick="sendChat('power readings')">Power Readings</span>
                    <span class="chat-suggest" onclick="sendChat('weather forecast')">Weather</span>
                    <span class="chat-suggest" onclick="sendChat('energy tips')">Energy Tips</span>
                    <span class="chat-suggest" onclick="sendChat('system overview')">System Overview</span>
                    <span class="chat-suggest" onclick="sendChat('help')">Help</span>
                </div>
                <div class="chat-input-row">
                    <input type="text" class="chat-input" id="chat-input" placeholder="Ask about your solar system..." onkeydown="if(event.key==='Enter')sendChat()">
                    <button class="chat-send" id="chat-send-btn" onclick="sendChat()">&#10148;</button>
                </div>
            </div>
        </div>

        <div id="timestamp" class="timestamp">Last Update: --</div>
        <footer><p>Copperbelt University, Kitwe, Zambia | Real-time ESP32 Data | Updates every second</p></footer>
    </div>

    <script>
        let powerChart;
        function initChart(){
            const ctx=document.getElementById('powerChart').getContext('2d');
            powerChart=new Chart(ctx,{type:'line',data:{labels:[],datasets:[
                {label:'Consumed (W)',data:[],borderColor:'#f44336',backgroundColor:'rgba(244,67,54,.1)',borderWidth:3,tension:.4,fill:true},
                {label:'Produced (W)',data:[],borderColor:'#4CAF50',backgroundColor:'rgba(76,175,80,.1)',borderWidth:3,tension:.4,fill:true}
            ]},options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{position:'top'},title:{display:true,text:'Last 60 Seconds'}},scales:{y:{beginAtZero:true,title:{display:true,text:'Watts'}},x:{title:{display:true,text:'Time'}}}}});
        }
        function sendRelayCmd(r,s){
            fetch('/api/relay',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({relay:r,state:s})})
            .then(r=>r.json()).then(d=>{
                console.log(d);
                alert('Relay '+r+' -> '+s+': '+d.status);
            }).catch(e=>{console.error(e);alert('Relay command failed: '+e);});
        }
        function testEmail(){
            const btn=document.getElementById('email-test-btn');
            const status=document.getElementById('email-status');
            btn.disabled=true;btn.textContent='Sending...';
            fetch('/api/test_email',{method:'POST'}).then(r=>r.json()).then(d=>{
                btn.disabled=false;btn.textContent='Send Test Email';
                status.innerHTML=d.status==='ok'?'<span style="color:#4CAF50">Email sent successfully!</span>':'<span style="color:#f44336">'+d.message+'</span>';
            }).catch(e=>{
                btn.disabled=false;btn.textContent='Send Test Email';
                status.innerHTML='<span style="color:#f44336">Error: '+e+'</span>';
            });
        }
        function exportCSV(){
            const s=document.getElementById('export-start').value||new Date(Date.now()-30*86400000).toISOString().slice(0,10);
            const e=document.getElementById('export-end').value||new Date().toISOString().slice(0,10);
            window.location='/api/export/csv?start='+s+'&end='+e+' 23:59:59';
        }
        function exportExcel(){
            const s=document.getElementById('export-start').value||new Date(Date.now()-30*86400000).toISOString().slice(0,10);
            const e=document.getElementById('export-end').value||new Date().toISOString().slice(0,10);
            window.location='/api/export/excel?start='+s+'&end='+e+' 23:59:59';
        }
        function exportPDF(){
            const m=document.getElementById('report-month').value;
            const y=document.getElementById('report-year').value;
            window.location='/api/report/monthly?month='+m+'&year='+y;
        }
        function fetchData(){
            fetch('/api/data').then(r=>r.json()).then(d=>{
                document.getElementById('battery-soc').innerHTML=d.battery_soc.toFixed(1)+'<span class="unit">%</span>';
                document.getElementById('battery-voltage').innerHTML=d.battery_voltage.toFixed(1)+'<span class="unit">V</span>';
                document.getElementById('battery-current').innerHTML=d.battery_current.toFixed(2)+'<span class="unit">A</span>';
                document.getElementById('load-power').innerHTML=d.load_power.toFixed(1)+'<span class="unit">W</span>';
                document.getElementById('inverter-power').innerHTML=d.inverter_power.toFixed(1)+'<span class="unit">W</span>';
                document.getElementById('load-current').innerHTML=d.load_current.toFixed(2)+'<span class="unit">A</span>';
                document.getElementById('energy-consumed').innerHTML=d.energy_consumed_kwh.toFixed(6)+'<span class="unit">kWh</span>';
                document.getElementById('energy-produced').innerHTML=d.energy_produced_kwh.toFixed(6)+'<span class="unit">kWh</span>';
                const b=document.getElementById('power-balance');
                b.innerHTML=d.power_balance.toFixed(1)+'<span class="unit">W</span>';
                b.className='value '+(d.power_balance>=0?'balance-positive':'balance-negative');
                document.getElementById('load1-badge').innerHTML='Load 1: '+d.load1_state;
                document.getElementById('load2-badge').innerHTML='Load 2: '+d.load2_state;
                document.getElementById('load1-badge').className='load-badge '+(d.load1_state==='ON'?'load-on':'load-off');
                document.getElementById('load2-badge').className='load-badge '+(d.load2_state==='ON'?'load-on':'load-off');
                document.getElementById('load-powers').innerHTML='L1: '+(d.load1_power||0).toFixed(0)+'W | L2: '+(d.load2_power||0).toFixed(0)+'W';
                const t=document.getElementById('trip-status');
                t.textContent=d.trip_state==='OVL'?'OVERLOAD':d.trip_state==='NOR'?'NORMAL':d.trip_state;
                t.className='status '+(d.trip_state==='OVL'?'status-overload':d.trip_state==='NOR'?'status-normal':'status-tripped');
                const e=document.getElementById('esp-status');
                e.innerHTML=d.esp32_online?'<span style="color:#4CAF50;font-weight:bold">ESP32 ONLINE</span> | '+d.last_esp32_seen:'<span style="color:#f44336;font-weight:bold">ESP32 OFFLINE</span>';
                if(d.predictions){
                    document.getElementById('pred-today').innerHTML=d.predictions.today_energy+'<span class="unit">kWh</span>';
                    document.getElementById('pred-tomorrow').innerHTML=d.predictions.tomorrow_energy+'<span class="unit">kWh</span>';
                    document.getElementById('pred-weekly').innerHTML=d.predictions.weekly_energy+'<span class="unit">kWh</span>';
                    const p=document.getElementById('peak-hours');
                    if(p&&d.predictions.peak_hours)p.innerHTML=d.predictions.peak_hours.map(x=>'<div class="peak-hour">'+x.hour+' ('+x.factor+'x)</div>').join('');
                }
                if(d.recommendations){
                    document.getElementById('recommendations').innerHTML=d.recommendations.map(r=>{
                        if(typeof r==='object'){
                            const sev=r.severity?'rec-'+r.severity:'';
                            return '<div class="recommendation '+sev+'"><span class="rec-icon">'+r.icon+'</span>'+r.text+'</div>';
                        }
                        return '<div class="recommendation">'+r+'</div>';
                    }).join('');
                }
                if(d.weather&&d.weather.daily&&d.weather.daily.length>0){const w=d.weather.daily[0];document.getElementById('weather-info').innerHTML='<div>Temp: '+w.temp_max+'C/'+w.temp_min+'C</div><div>Condition: '+d.weather.current.condition+'</div>';}
                if(d.sunlight)document.getElementById('sunlight-info').innerHTML='<div>Sunrise: '+d.sunlight.sunrise+'</div><div>Sunset: '+d.sunlight.sunset+'</div><div>UV: '+d.sunlight.uv_index+'</div>';
                document.getElementById('timestamp').innerHTML='Last Update: '+d.timestamp;
            }).catch(e=>console.error('Fetch data error:',e));
        }
        function fetchRealTimePower(){
            fetch('/api/real_time_power').then(r=>r.json()).then(d=>{
                if(powerChart&&d.consumed&&d.produced){
                    powerChart.data.labels=d.timestamps.slice(-30);
                    powerChart.data.datasets[0].data=d.consumed.slice(-30);
                    powerChart.data.datasets[1].data=d.produced.slice(-30);
                    powerChart.update();
                }
            }).catch(e=>console.error('Realtime error:',e));
        }
        function loadHistoricalChart(){
            fetch('/api/historical_data').then(r=>r.json()).then(d=>{
                if(d.timestamps&&d.timestamps.length>0){
                    Plotly.newPlot('historicalChart',[
                        {x:d.timestamps,y:d.consumed,name:'Consumed (W)',type:'scatter',mode:'lines',line:{color:'#f44336',width:3}},
                        {x:d.timestamps,y:d.produced,name:'Produced (W)',type:'scatter',mode:'lines',line:{color:'#4CAF50',width:3}}
                    ],{title:'2-Month Power History',xaxis:{title:'Date',tickangle:-45},yaxis:{title:'Power (W)'},height:420,hovermode:'closest'},{responsive:true});
                } else {
                    document.getElementById('historicalChart').innerHTML='<div style="text-align:center;padding:40px;color:#999">No historical data yet. Data will appear after ESP32 sends readings.</div>';
                }
            }).catch(e=>{console.error('Historical error:',e);document.getElementById('historicalChart').innerHTML='<div style="text-align:center;padding:40px;color:#f44336">Error loading historical data</div>';});
        }
        function loadComparisonChart(){
            fetch('/api/comparison').then(r=>r.json()).then(d=>{
                const traces=[];
                if(d.last_month&&d.last_month.dates&&d.last_month.dates.length>0){
                    traces.push({x:d.last_month.dates,y:d.last_month.consumed,name:d.last_month_label+' Consumed',type:'scatter',mode:'lines',line:{color:'#ff9800',width:2,dash:'dash'}});
                    traces.push({x:d.last_month.dates,y:d.last_month.produced,name:d.last_month_label+' Produced',type:'scatter',mode:'lines',line:{color:'#9e9e9e',width:2,dash:'dash'}});
                }
                if(d.this_month&&d.this_month.dates&&d.this_month.dates.length>0){
                    traces.push({x:d.this_month.dates,y:d.this_month.consumed,name:d.this_month_label+' Consumed',type:'scatter',mode:'lines',line:{color:'#f44336',width:3}});
                    traces.push({x:d.this_month.dates,y:d.this_month.produced,name:d.this_month_label+' Produced',type:'scatter',mode:'lines',line:{color:'#4CAF50',width:3}});
                }
                if(traces.length===0){
                    document.getElementById('comparisonChart').innerHTML='<div style="text-align:center;padding:40px;color:#999">No comparison data yet. Data will appear after ESP32 sends readings.</div>';
                } else {
                    Plotly.newPlot('comparisonChart',traces,{title:'This Month vs Last Month',xaxis:{title:'Date'},yaxis:{title:'Power (W)'},height:450,hovermode:'closest'},{responsive:true});
                }
            }).catch(e=>{console.error('Comparison error:',e);document.getElementById('comparisonChart').innerHTML='<div style="text-align:center;padding:40px;color:#f44336">Error loading comparison data</div>';});
        }
        function sendChat(msg){
            const input=document.getElementById('chat-input');
            const messages=document.getElementById('chat-messages');
            const text=msg||input.value.trim();
            if(!text)return;
            input.value='';
            messages.innerHTML+='<div class="chat-msg user"><div class="chat-avatar user-avatar">U</div><div class="chat-bubble user">'+text.replace(/</g,'&lt;')+'</div></div>';
            messages.scrollTop=messages.scrollHeight;
            const sendBtn=document.getElementById('chat-send-btn');
            sendBtn.disabled=true;
            messages.innerHTML+='<div class="chat-msg" id="chat-typing"><div class="chat-avatar bot-avatar">AI</div><div class="chat-bubble bot" style="opacity:.6">Thinking...</div></div>';
            messages.scrollTop=messages.scrollHeight;
            fetch('/api/chat',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({message:text})})
            .then(r=>r.json()).then(d=>{
                const typing=document.getElementById('chat-typing');
                if(typing)typing.remove();
                messages.innerHTML+='<div class="chat-msg"><div class="chat-avatar bot-avatar">AI</div><div class="chat-bubble bot">'+d.response.replace(/\\n/g,'<br>').replace(/</g,'&lt;').replace(/&lt;br>/g,'<br>')+'</div></div>';
                messages.scrollTop=messages.scrollHeight;
                sendBtn.disabled=false;
            }).catch(e=>{
                const typing=document.getElementById('chat-typing');
                if(typing)typing.remove();
                messages.innerHTML+='<div class="chat-msg"><div class="chat-avatar bot-avatar">AI</div><div class="chat-bubble bot" style="background:#f44336">Error: Could not reach AI assistant.</div></div>';
                messages.scrollTop=messages.scrollHeight;
                sendBtn.disabled=false;
            });
        }
        const d=new Date();document.getElementById('export-start').value=new Date(d-30*86400000).toISOString().slice(0,10);document.getElementById('export-end').value=d.toISOString().slice(0,10);document.getElementById('report-month').value=d.getMonth()+1;
        initChart();fetchData();fetchRealTimePower();loadHistoricalChart();loadComparisonChart();
        setInterval(fetchData,1000);setInterval(fetchRealTimePower,2000);setInterval(loadComparisonChart,60000);
    </script>
</body>
</html>"""

# ========== INIT (runs on import - works with gunicorn) ==========
print("=" * 70)
print("SOLAR MICROGRID ENERGY MANAGEMENT SYSTEM")
print("=" * 70)
init_db()
if not os.path.exists(HISTORICAL_DATA_FILE):
    init_historical_data()
else:
    print("Loading historical data...")
    try:
        df = pd.read_csv(HISTORICAL_DATA_FILE)
        df['timestamp'] = pd.to_datetime(df['timestamp'])
        for _, row in df.iterrows():
            historical_power.append({'timestamp': row['timestamp'], 'consumed': row['consumed_power'], 'produced': row['produced_power']})
        print(f"Loaded {len(historical_power)} records")
    except Exception as e:
        print(f"Error: {e}")
        init_historical_data()
get_weather_data()
threading.Thread(target=update_weather_periodically, daemon=True).start()
threading.Thread(target=check_esp32_status, daemon=True).start()
threading.Thread(target=check_battery_email, daemon=True).start()
print("\nSystem Ready! Dashboard: http://localhost:5000")

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=False, use_reloader=False)
